from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.select import Select
import time
from openpyxl import Workbook, load_workbook
import datetime
import os

c_date = datetime.date.today()  # current_date
# starting creating excel file
try:
    wb = load_workbook(f'canvas_champ_data_{c_date}.xlsx')
    active_sheet = wb['Sheet']
except FileNotFoundError:
    wb = Workbook()
    active_sheet = wb['Sheet']
    active_sheet.append(['Website', 'Date', 'Height', 'Width', 'Discounted Price', 'Actual Price', 'Discount Percent',
                         'Standard Shipping Price', 'Express Shipping Price', 'Priority Shipping Price'])
    wb.save(f'canvas_champ_data_{c_date}.xlsx')
except:
    print('something went wrong')

site_url = "https://canvaschamp.com/"

browser = webdriver.Chrome('C:\\Users\\thaku\\Desktop\\inte\\chromedriver.exe')
browser.maximize_window()
browser.get(site_url)

time.sleep(3)

size_select_div = browser.find_element(By.XPATH,
                                       '//li[@class="size_selectbox clearfix"]')  # custom canvas size main div
custom_size_checkbox = browser.find_element(By.XPATH,
                                            '//li[@class="size_selectbox clearfix"]//input[@name="name_size"]')
custom_size_checkbox.click()  # clicking the custome size check box
time.sleep(5)
height_dropdown = size_select_div.find_element(By.XPATH, '//select[@name="canvas_height"]')
width_dropdown = size_select_div.find_element(By.XPATH, '//select[@name="canvas_width"]')
height_list = []
width_list = []
for option in Select(height_dropdown).options:
    height_list.append(option.text)

for option in Select(width_dropdown).options:
    width_list.append(option.text)

size_combinations = []
for height_op in height_list:
    for width_op in width_list:
        single_comb = []
        single_comb.append(height_op)
        single_comb.append(width_op)
        size_combinations.append(single_comb)
print(size_combinations)
browser.quit()

for combination in size_combinations:


    try:

        print(combination)

        browser = webdriver.Chrome('C:\\Users\\thaku\\Desktop\\inte\\chromedriver.exe')
        browser.maximize_window()
        browser.get(site_url)

        data_list = []
        # canvas_type = "Not Available"
        data_list.append(site_url)
        data_list.append(datetime.date.today())

        Select(browser.find_element(By.XPATH, '//select[@name="canvas_height"]')).select_by_visible_text(combination[0])
        time.sleep(1)
        Select(browser.find_element(By.XPATH, '//select[@name="canvas_width"]')).select_by_visible_text(combination[1])
        time.sleep(1)

        data_list.append(
            Select(browser.find_element(By.XPATH, '//select[@name="canvas_height"]')).first_selected_option.text)
        time.sleep(1)
        data_list.append(
            Select(browser.find_element(By.XPATH, '//select[@name="canvas_width"]')).first_selected_option.text)

        time.sleep(3)

        data_list.append(
            browser.find_elements(By.XPATH, '//div[@class="price_box"]')[-1].find_element(By.TAG_NAME, 'span').text)

        data_list.append(
            browser.find_elements(By.XPATH, '//div[@class="price_box"]')[-1].find_element(By.TAG_NAME, 'strike').text)
        data_list.append(browser.find_elements(By.XPATH, '//div[@class="price_box"]')[-1].find_element(By.CLASS_NAME,
                                                                                                       'offer_label').text)
        # data_list.append(canvas_type)
        # data_list.append('Not Applicable')

        time.sleep(2)

        browser.find_element(By.XPATH, "//button[contains(text(),'Start Order' )]").click()

        time.sleep(30)

        browser.switch_to.frame(browser.find_element_by_xpath("//iframe[@id='desgntool-iframe']"))

        browser.find_element_by_xpath("//a[@currenttab='upload_photo']")
        # s_browser = browser.find_element_by_id("maincontent")
        print("===========================")
        upload = browser.find_element_by_xpath("//input[contains(@name,'uploadart[]')]").send_keys(
            os.getcwd() + "/tile.jpg")

        time.sleep(10)

        button = browser.find_element_by_class_name('addtocart')
        button.click()

        time.sleep(15)

        # standard shipping
        standard_ship = browser.find_element_by_xpath("//label[contains(text(),'Standard')]")
        data_list.append(standard_ship.text.split()[1])  # price

        time.sleep(4)
        # Express shipping
        express_ship = browser.find_element_by_xpath("//label[contains(text(),'Express')]")
        data_list.append(express_ship.text.split()[1])

        time.sleep(4)
        # Priority shiping
        prior_ship = browser.find_element_by_xpath("//label[contains(text(),'Priority')]")
        data_list.append(prior_ship.text.split()[1])

        browser.switch_to.default_content()

        if data_list not in active_sheet:
            active_sheet.append(data_list)
            wb.save(f'canvas_champ_data_{c_date}.xlsx')

            print(data_list)

        browser.quit()
        time.sleep(2)

    except:
        

        with open("canvas_champ_missing.txt", "a") as file1:
            file1.write(f"{datetime.date.today()} error occure in {combination}")

        print('error occure in', combination)
        browser.quit()
        time.sleep(2)


